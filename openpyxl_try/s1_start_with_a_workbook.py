from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""
create a new workbook/worksheet
"""
# wb = Workbook()
# file_name = 'try_wb.xlsx'
#
# ws1 = wb.active
# ws1.title = 'synferlo'
# # ws fill in the content row by row, like dataframe.
# for row in range(1,50):
#     ws1.append(range(600))
# # create a new worksheet
# ws2 = wb.create_sheet(title = 'yan')
#
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title = 'adam')
# # fill in the content in particular row and col by using a for loop and ws.cell(column = , row = , value = )
# for row in range(5,10):
#     for col in range(2,5):
#         content = ws3.cell(column = col, row = row, value = '{0}'.format(get_column_letter(col)))
#
# wb.save(filename = file_name)



"""
load a existing xlsx file
"""
# from openpyxl import load_workbook
#
# wb = load_workbook('try_wb.xlsx')
# print(wb.sheetnames)
# #['synferlo', 'yan', 'adam']
#
# sheet_range = wb['adam']
# print(sheet_range)
# #<Worksheet "adam">



"""
add date/time to the xlsx
"""
# from openpyxl import load_workbook
# import datetime
#
# wb = load_workbook('try_wb.xlsx')
# ws = wb.create_sheet('datetime')
# ws['A1'] = datetime.datetime(2020,5,2,20,59,00)
# print(ws['A1'].value)
# #2020-05-02 20:59:00
#
# wb.save('try_wb.xlsx')



"""
to delete sheets in a workbook
"""
# from openpyxl import load_workbook
#
#
# wb = load_workbook('try_wb.xlsx')
# print(wb.sheetnames)
#
# sheet_name = wb.sheetnames
## delete a content in the list
# sheet_name.remove('yan')
# print(sheet_name)
#
# for i in sheet_name:
#     del wb[i]
#
# print(wb.sheetnames)
# wb.save('try_wb.xlsx')


"""

"""