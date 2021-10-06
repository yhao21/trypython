from pydocx import PyDocX
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
import glob, re, os
import pandas as pd


def to_html():

    #transfer all docx files into html through this for loop.
    for one_file in glob.glob('*.docx'):
        #extract the name of each employee
        name = re.compile(r'(.*?)员工履历').findall(one_file)[0]
        #extract content of docx file
        html = PyDocX.to_html(one_file)
        #create a html file
        f = open(name + '.html', 'w', encoding = 'utf-8')
        #save the content into this html file.
        f.write(html)
        #save and close the html file.
        f.close()


def to_excel(writer):

    #transfer each html file in to one sheet in workbook.
    for file in glob.glob('*.html'):
        #extract the name of each employee.
        sheet_name = file.replace('.html', '')
        #extract the content of html file. It should contain 6 sub-tables, as there are 6 tables in docx file.
        df = pd.read_html(file, encoding = 'utf-8')
        #edit the first sub-table
        table0 = pd.DataFrame(df[0])
        table0.iloc[4, -1] = ''
        table0.iloc[5, -1] = ''
        #edit the rests
        for i in range(2, 7):
            table0.iloc[6, i] = ''
        #create a new dataframe, we will put all those 6 sub-tables in 1 dataframe.
        data = pd.DataFrame()
        #create a dataframe with only one row. we will separate each sub-table by a blank row.
        blank = pd.DataFrame([''])

        for i in range(1, 6):
            #transfer sub-table2-6 into dataframe. we have already transfered the first sub-table into dataframe in line 33.
            table = pd.DataFrame(df[i])
            for i in range(1, table.shape[1]):
                #delete the overlapping value, title, in each subtable.
                table.iloc[0, i] = ''
            #combine sub-table2-6 into one dataframe.
            data = pd.concat([data, blank, table], axis = 0, ignore_index = True)
        #combine the first sub-table with the rest sub-tables into one dataframe. this dataframe contains all the information for each employee.
        data = pd.concat([table0, data], axis = 0, ignore_index = True)
        #transfer this dataframe to one sheet, named by the name of this employee.
        #each round in this for loop will create a new sheet which contains all the information for one employee.
        data.to_excel(writer, index = False, sheet_name = sheet_name)
        #delete the html file we have just parsed. save the space of your disk.
        os.remove(file)


def cell_merge(ws):

    #define the cells need to merge in table1 (info of employee)
    cell_table1 = ['F' + str(i) + ':' + 'H' + str(i) for i in range(1, 7)] + ['B7:H7']
    #gather the cells contain the title of each subtable
    title_cell_num = []
    #define what is the value of each title
    title_item = ['任职信息', '绩效考核', '学历信息', '内部工作履历（最近5条）', '外部工作履历']
    #find the location of each title, save into 'title_cell_num' (a list).
    for i in range(1, 50):
        for j in title_item:
            if ws['A' + str(i)].value == j:
                row_info = i
                title_cell_num.append(row_info)
    #define the range of cells, with title, need to merge.
    cell_title = ['A' + str(i) + ":" + 'H' + str(i) for i in title_cell_num]
    #define the cells need to merger in sub-table2.
    table2_rows = [title_cell_num[0] + i for i in range(1, 6)]
    cell_table2 = ['F' + str(i) + ':' + 'H' + str(i) for i in table2_rows]
    #define the cells need to merger in sub-table5.
    table5_rows = [title_cell_num[-2] + i for i in range(1, 7)]
    cell_table5 = ['F' + str(i) + ':' + 'H' + str(i) for i in table5_rows]
    #define the cells need to merger in sub-table6.
    table6_rows = [title_cell_num[-1] + i for i in range(1, 7)]
    cell_table6 = ['E' + str(i) + ':' + 'H' + str(i) for i in table6_rows]
    #compile all target cells into one list.
    all_merge_cell = cell_table1 + cell_title + cell_table2 + cell_table5 + cell_table6
    # merge.
    for cell in all_merge_cell:
        ws.merge_cells(cell)
    #bold the title.
    cell_title_after_merge = ['A' + str(i) for i in title_cell_num]
    for cell in cell_title_after_merge:
        ws[cell].font = Font(bold = True)

    return title_cell_num


def setup_border(col_lim, row_lim, ws, line_style):

    #border each cell.
    for col in col_lim:
        for row in row_lim:
            ws[col + str(row)].border = Border(top = line_style, bottom = line_style, left = line_style, right = line_style)


def prepare_border(ws, line, title):
    #prepare the rows and columns to be bordered.
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    row_table1 = [i for i in range(1, 8)]
    row_table2 = [i for i in range(9, 15)]
    row_table3 = [i for i in range(16, 19)]
    #as individual may have different numbers of info for education background, we need to know the range of rows for this sub-table.
    for i in range(8):
        if ws['A' + str(title[2] + i)].value == None:
            end_row_table4 = title[2] + i
            break
    row_table4 = [i for i in range(title[2], end_row_table4)]
    row_table5 = [i for i in range(title[3], title[3] + 7)]
    row_table6 = [i for i in range(title[4], title[4] + 7)]
    #compile all the rows need to be bordered.
    border_rows = row_table1 + row_table2 + row_table3 + row_table4 + row_table5 + row_table6
    #use this function to border cells.
    setup_border(row_lim = border_rows, col_lim = columns, ws = ws, line_style = line)


def color_fill(ws,color_cells):

    #color each cell.
    for cell in color_cells:
        ws[cell].fill = PatternFill(fill_type = 'solid', fgColor = 'AAAAAA')


def pre_color_fill(ws, title):
    #compile all cells need to be greyed
    columns = ['A','C','E']
    full_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    color_table1 = [col + str(row) for col in columns for row in range(1,7)] + ['A7']
    color_table2 = [col + str(row) for col in columns for row in range(10,15)]
    color_table3_6 = [col + str(title[i] + 1) for col in full_col for i in range(1,5)]
    color_cells = color_table1 + color_table2 + color_table3_6
    #get white background
    for i in range(1,100):
        for j in range(1,100):
            ws.cell(row = i, column = j).fill = PatternFill(fill_type = 'solid', fgColor = 'FFFFFF')
    #grey the cells
    color_fill(ws, color_cells)


def edit_excel(excel_name):

    #open the excel contains all the information of each employee. the info is gathered from each html file. we have done this part in line 58 (to_excel).
    wb = load_workbook(excel_name)
    #edit each sheet by using a for loop.
    for sheet in wb.sheetnames:
        #open the target sheet.
        ws = wb[sheet]
        #prepare the line style of the border.
        thin = Side(border_style = 'thin', color = '000000')
        #delete the indexing row in each sheet.
        ws.delete_rows(1)
        #merge cells
        title_row = cell_merge(ws)
        #border cells
        prepare_border(ws = ws, line = thin, title = title_row)
        #color cells
        pre_color_fill(ws = ws, title = title_row)
    #save workbook after editing each sheet.
    wb.save(excel_name)


if __name__ == '__main__':
    #create a new excel file.
    excel_name = 'employee_data.xlsx'
    writer = pd.ExcelWriter(excel_name)
    #transfer all docx file into html file.
    to_html()
    #parse all html files, save data in each html to a sheet.
    to_excel(writer)
    writer.save()
    #merge, border, color the cells.
    edit_excel(excel_name)


