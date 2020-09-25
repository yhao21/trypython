from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font,Border,Side,PatternFill



def set_border(col_lim,row_lim,line):

    for col in col_lim:
        for row in row_lim:
            location = col + str(row)
            target_cell = ws[location]
            target_cell.border = Border(top=line, bottom=line, left=line, right=line)


#合并单元格
def cell_merge():

    sub_cell = ['E' + str(i) + ':' + 'F' + str(i) for i in range(33, 39)]
    cell_merge = ['A9:F9','A16:H16','A20:H20','A24:F24','A32:F32','B7:F7'] + sub_cell
    for cell in cell_merge:
        ws.merge_cells(cell)

    cells = ['A9','A16','A20','A24','A32']
    cell_outer = ['E33','E34','E35','E36','E37','E38']
    for cell in cells:
        # ws[cell].alignment = Alignment(horizontal='center',vertical='center')
        ws[cell].font = Font(bold = True)
    # for cell in cell_outer:
    #     ws[cell].alignment = Alignment(horizontal='center',vertical='center')

#设置边框
def setup_border(line):

    col_short = ['A','B','C','D','E','F']
    col_long = col_short + ['G','H']
    row_info = [i for i in range(1,8)]
    row_occ = [i for i in range(9,15)]
    row_kpi = [i for i in range(16,19)]
    row_edu = [i for i in range(20,23)]
    row_inner = [i for i in range(24,31)]
    row_outer = [i for i in range(32,39)]
    set_border(col_lim = col_short, row_lim = row_info, line = line)
    set_border(col_lim = col_short, row_lim = row_occ, line = line)
    set_border(col_lim = col_long, row_lim = row_kpi, line = line)
    set_border(col_lim = col_long, row_lim = row_edu, line = line)
    set_border(col_lim = col_short, row_lim = row_inner, line = line)
    set_border(col_lim = col_short, row_lim = row_outer, line = line)

def color_fill(sheet,color_cells):

    for cell in color_cells:
        sheet[cell].fill = PatternFill(fill_type='solid',fgColor='AAAAAA')


if __name__ == '__main__':
    wb = load_workbook('test.xlsx')
    for sheet in wb.sheetnames:

        ws = wb[sheet]
        thin = Side(border_style='thin', color='000000')
        ws.delete_rows(1)
        cell_merge()
        setup_border(line = thin)
        table_info_cell = ['A7']
        table_occ_cell = []
        table_kpi_cell = [j + '17' for j in ['A','B','C','D','E','F','G','H']]
        table_edu_cell = [j + '21' for j in ['A','B','C','D','E','F','G','H']]
        table_inner_cell = [j + '25' for j in ['A','B','C','D','E','F']]
        table_outer_cell = [j + '33' for j in ['A','B','C','D','E','F']]
        for i in range(1,7):
            for j in ['A','C','E']:
                x = j + str(i)
                table_info_cell.append(x)
        for i in range(10,15):
            for j in ['A','C','E']:
                x = j + str(i)
                table_occ_cell.append(x)

        color_cells = table_info_cell+table_occ_cell+table_kpi_cell+table_edu_cell+table_inner_cell+table_outer_cell
        color_fill(ws,color_cells)

    wb.save('new_test.xlsx')